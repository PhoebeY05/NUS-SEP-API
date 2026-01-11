import os
import re
import unicodedata

import pandas as pd
from openpyxl import load_workbook

# ========= CONFIG =========
EXCEL_PATH = "SEP-Credit-Transfer-Calculator.xlsx"
OUTPUT_DIR = "data"

os.makedirs(OUTPUT_DIR, exist_ok=True)

# ========= STEP 1: LOAD WORKBOOK =========
wb = load_workbook(EXCEL_PATH, data_only=True)

print("Scanning workbook...")

# ========= STEP 2: EXTRACT HIDDEN SHEETS =========
hidden_csv_paths = {}

for ws in wb.worksheets:
    if ws.sheet_state == "hidden":
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(row)

        df = pd.DataFrame(data)

        csv_name = ws.title.replace(" ", "_") + ".csv"
        csv_path = os.path.join(OUTPUT_DIR, csv_name)

        df.to_csv(csv_path, index=False, header=False)
        hidden_csv_paths[ws.title] = csv_path

        print(f"✔ Extracted hidden sheet: {ws.title} → {csv_path}")

# ========= STEP 3: LOAD CREDIT SCORES =========
if "Credit Scores" not in hidden_csv_paths:
    raise RuntimeError("Credit Scores sheet not found!")

credit_scores_path = hidden_csv_paths["Credit Scores"]

df = pd.read_csv(credit_scores_path, header=None)

# ========= STEP 4: ASSIGN COLUMN NAMES =========
df.columns = [
    "region",
    "faculty",
    "country",
    "university",
    "partner_credit_value",
    "unit_type",
    "conversion_factor",
    "notes"
]

print("✔ Parsed Credit Scores table")

# ========= STEP 5: EXPORT CLEAN UNIVERSITY LIST =========
universities = (
    df[["region", "country", "university"]]
    .dropna()
    .drop_duplicates()
    .sort_values(["region", "country", "university"])
)

universities_path = os.path.join(OUTPUT_DIR, "universities.csv")
universities.to_csv(universities_path, index=False)

print(f"✔ Exported university list → {universities_path}")

# ========= STEP 5.5: MAP ORGANISATION IDs =========
def _normalize_text(s: str) -> str:
    if not isinstance(s, str):
        s = str(s)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.lower().strip()
    s = re.sub(r"[\-\u2013\u2014]", " ", s)  # hyphens to space
    s = re.sub(r"&", " and ", s)
    s = re.sub(r"[^a-z0-9 ]", "", s)  # drop punctuation
    s = re.sub(r"\s+", " ", s)
    # common abbreviation normalizations
    s = s.replace("univ", "university").replace(" uni ", " university ")
    s = s.replace("inst", "institute").replace(" sch ", " school ")
    return s.strip()

def _interactive_fuzzy_fill_ids(uni_df: pd.DataFrame, org_df: pd.DataFrame, threshold: float = 0.86) -> pd.DataFrame:
    """Interactively suggest organisation IDs for unmatched university names.

    - Shows top candidates by difflib similarity on normalized names.
    - Accept by entering option number; 'n' for next page; 's' skip; 'q' quit;
      or 'id:<VALUE>' to set an ID manually.
    - Can search in organisations.csv to manually set ID
    """
    import difflib

    uni_df = uni_df.copy()
    org_df = org_df.copy()

    # Normalized views
    uni_df["_norm"] = uni_df["university"].astype(str).map(_normalize_text)
    org_df["_norm"] = org_df["name"].astype(str).map(_normalize_text)

    # Build candidate arrays once
    org_norms = org_df["_norm"].tolist()
    org_ids = org_df["id"].tolist()
    org_names = org_df["name"].tolist()

    unmatched_idx = uni_df[(uni_df["id"].isna()) | (uni_df["id"].astype(str).str.strip() == "")].index
    if len(unmatched_idx) == 0:
        print("ℹ No unmatched universities for fuzzy matching.")
        return uni_df.drop(columns=["_norm"], errors="ignore")

    print(f"🔎 Starting interactive fuzzy matching for {len(unmatched_idx)} unmatched entries.")
    print("Enter option number to accept; 'n' next page; 's' skip; 'q' quit; or 'id:<VALUE>' for manual ID.")

    for idx in unmatched_idx:
        u_name = uni_df.at[idx, "university"]
        u_norm = uni_df.at[idx, "_norm"]
        # Compute scores
        scores = []
        for on, oid, oraw in zip(org_norms, org_ids, org_names):
            r = difflib.SequenceMatcher(None, u_norm, on).ratio()
            scores.append((r, oid, oraw))
        scores.sort(key=lambda x: x[0], reverse=True)

        page = 0
        while True:
            start = page * 5
            end = min(start + 5, len(scores))
            print("\nUniversity:", u_name)
            print("Candidates (", start + 1, "-", end, "):")
            for i in range(start, end):
                r, oid, oraw = scores[i]
                flag = "✅" if r >= threshold else "  "
                print(f"  {i+1}. {oraw}  [id={oid}]  similarity={r:.3f} {flag}")

            choice = input("Choice: ").strip()
            if choice.lower() == "q":
                print("⏹ Quit fuzzy matching.")
                return uni_df.drop(columns=["_norm"], errors="ignore")
            if choice.lower() == "s":
                break
            if choice.lower() == "n":
                page = min(page + 1, max(0, (len(scores) - 1) // 5))
                continue
            if choice.lower().startswith("id:"):
                manual_id = choice.split(":", 1)[1].strip()
                uni_df.at[idx, "id"] = manual_id
                print(f"✔ Set manual ID {manual_id} for '{u_name}'.")
                break
            if choice.isdigit():
                sel = int(choice)
                if 1 <= sel <= len(scores):
                    r, oid, oraw = scores[sel - 1]
                    if r < threshold:
                        confirm = input(f"Similarity {r:.3f} below threshold {threshold}. Accept? [y/N]: ").strip().lower()
                        if confirm != "y":
                            continue
                    uni_df.at[idx, "id"] = oid
                    print(f"✔ Mapped '{u_name}' → '{oraw}' (id={oid}, sim={r:.3f}).")
                    break
                else:
                    print("✖ Invalid selection.")
            else:
                print("✖ Invalid input.")

    return uni_df.drop(columns=["_norm"], errors="ignore")

try:
    org_path = os.path.join(OUTPUT_DIR, "organisations.csv")
    if os.path.exists(org_path):
        org_df = pd.read_csv(org_path)
        if {"id", "name"}.issubset(org_df.columns):
            org_df["_key"] = org_df["name"].astype(str).map(_normalize_text)
            uni_df = universities.copy()
            uni_df["_key"] = uni_df["university"].astype(str).map(_normalize_text)
            id_map = dict(zip(org_df["_key"], org_df["id"]))
            uni_df["id"] = uni_df["_key"].map(id_map)
            # Interactive fuzzy mapping
            # Use a clean org_df (without _key) for display
            org_view = org_df.drop(columns=["_key"], errors="ignore")
            uni_df = _interactive_fuzzy_fill_ids(uni_df, org_view)
            # Cleanup and write back
            uni_df = uni_df.drop(columns=["_key"], errors="ignore")
            uni_df.to_csv(universities_path, index=False)
            # Run validation after IDs are persisted to CSV
            print(f"✔ Mapped organisation IDs → {universities_path}" + (" (fuzzy confirmed)" if fuzzy_flag else ""))
        else:
            print("ℹ organisations.csv missing 'id'/'name' columns; skipping ID map")
    else:
        print("ℹ organisations.csv not found; skipping ID map")
except Exception as e:
    print(f"⚠ Failed to map organisation IDs: {e}")

# ========= STEP 6: OPTIONAL DERIVED DATASETS =========

# Universities by faculty
universities_by_faculty = (
    df[["faculty", "country", "university"]]
    .dropna()
    .drop_duplicates()
    .sort_values(["faculty", "country", "university"])
)

universities_by_faculty_path = os.path.join(
    OUTPUT_DIR, "universities_by_faculty.csv"
)
universities_by_faculty.to_csv(universities_by_faculty_path, index=False)

print(f"✔ Exported universities by faculty → {universities_by_faculty_path}")

# ========= STEP 7: OPTIONAL HELPER FUNCTION =========
def get_universities(faculty: str):
    return (
        df[df["faculty"] == faculty]["university"]
        .dropna()
        .unique()
        .tolist()
    )

# ======== STEP 8: POST-PROCESS IDs (validate, combine, prompt) =========
def _report_non_school_ids(uni_df_check: pd.DataFrame, org_df_check: pd.DataFrame) -> None:
    try:
        view = org_df_check[["id", "org", "name"]].copy()
    except Exception:
        print("ℹ organisations.csv missing expected columns for validation; skipping org type check")
        return
    merged = uni_df_check.merge(view, on="id", how="left")
    # Treat NaN IDs as empty
    has_id = merged["id"].notna() & merged["id"].astype(str).str.strip().ne("")
    not_school = merged["org"].astype(str).str.upper().ne("SCHL")
    bad = merged[has_id & not_school]
    count = len(bad)
    if count == 0:
        print("✔ All mapped IDs belong to organisations of type SCHL")
        return
    print(f"⚠ {count} mapped IDs belong to organisations not marked as SCHL.")
    sample = bad[["university", "id", "org", "name"]]
    print("All:")
    for _, row in sample.iterrows():
        print(f"  - university='{row['university']}', id={row['id']}, org={row['org']}, org_name='{row['name']}'")

def _combine_identical_ids_interactive(uni_df: pd.DataFrame, org_df: pd.DataFrame) -> pd.DataFrame:
    """Interactively collapse duplicate IDs by choosing a canonical name per ID.

    For each ID with multiple rows, shows existing `university` names and the
    organisation's official name, and prompts for which to keep.
    """
    uni_df = uni_df.copy()
    # Consider NaN as empty (no ID)
    has_id_mask = uni_df["id"].notna() & uni_df["id"].astype(str).str.strip().ne("")
    with_ids = uni_df[has_id_mask].copy()
    without_ids = uni_df[~has_id_mask].copy()

    org_name_map = dict(zip(org_df.get("id", []), org_df.get("name", [])))

    collapsed_rows: list[pd.Series] = []
    grouped = with_ids.groupby("id", sort=False)
    total_dupe_groups = 0
    for gid, group in grouped:
        if len(group) == 1:
            collapsed_rows.append(group.iloc[0])
            continue
        total_dupe_groups += 1
        current_names = group["university"].astype(str).tolist()
        org_official = org_name_map.get(gid, None)

        print("\nDuplicate ID detected:")
        print(f"  id={gid}")
        if org_official:
            print(f"  organisation official name: {org_official}")
        print("  candidate university names:")
        for i, nm in enumerate(current_names, start=1):
            print(f"    {i}. {nm}")
        print("  o. Use organisation official name" + (f" ('{org_official}')" if org_official else " (unavailable)"))
        print("  s. Skip (keep first)\n")

        while True:
            choice = input("Choose canonical name [number/o/s]: ").strip().lower()
            if choice == "s" or choice == "":
                chosen_name = current_names[0]
                break
            if choice == "o":
                if org_official:
                    chosen_name = org_official
                else:
                    print("✖ Organisation name unavailable; select a number.")
                    continue
                break
            if choice.isdigit():
                idx = int(choice)
                if 1 <= idx <= len(current_names):
                    chosen_name = current_names[idx - 1]
                    break
                else:
                    print("✖ Invalid selection.")
            else:
                print("✖ Invalid input.")

        # Keep the first row's other fields, set canonical name
        base = group.iloc[0].copy()
        base["university"] = chosen_name
        collapsed_rows.append(base)

    if total_dupe_groups:
        print(f"ℹ Collapsed {total_dupe_groups} duplicate ID groups interactively")

    combined = pd.concat([pd.DataFrame(collapsed_rows), without_ids], ignore_index=True)
    return combined

try:
    org_path = os.path.join(OUTPUT_DIR, "organisations.csv")
    uni_path = os.path.join(OUTPUT_DIR, "universities.csv")
    if os.path.exists(org_path) and os.path.exists(uni_path):
        org_df_final = pd.read_csv(org_path)
        uni_df_final = pd.read_csv(uni_path)
        print("\n# Step 8: ID post-processing")
        # 1) Check for non-SCHL ids
        _report_non_school_ids(uni_df_final, org_df_final)
        # 2) Combine identical ids (interactive)
        uni_df_final = _combine_identical_ids_interactive(uni_df_final, org_df_final)
        # 3) Prompt for still empty ids (interactive)
        # Treat NaN or empty string as empty ID
        empty_mask = uni_df_final["id"].isna() | uni_df_final["id"].astype(str).str.strip().eq("")
        if empty_mask.any():
            print(f"ℹ {empty_mask.sum()} entries still have empty IDs — starting interactive prompt.")
            # Use only necessary columns from orgs
            org_view_final = org_df_final[["id", "name"]].copy()
            uni_df_final = _interactive_fuzzy_fill_ids(uni_df_final, org_view_final)
        else:
            print("✔ No empty IDs remain.")
        # Write back
        uni_df_final.to_csv(uni_path, index=False)
        print(f"✔ Step 8 complete → {uni_path}")
    else:
        print("ℹ Skipping Step 8: universities.csv or organisations.csv not found")
except Exception as e:
    print(f"⚠ Step 8 failed: {e}")

# ======== STEP 9: CLEANUP =========
# Files to DELETE
DELETE_FILES = {
    # "universities_by_country.csv",
    # "universities_by_faculty.csv",
    # "Credit_Scores.csv",
    "List.csv"
}

print("Starting cleanup...")

for filename in os.listdir(OUTPUT_DIR):
    file_path = os.path.join(OUTPUT_DIR, filename)

    # Only touch files (ignore folders)
    if os.path.isfile(file_path) and filename in DELETE_FILES:
        os.remove(file_path)
        print(f"🗑 Deleted: {filename}")

print("\n✅ Cleanup complete.")
print("Remaining files:", os.listdir(OUTPUT_DIR))